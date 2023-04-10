import matplotlib.pyplot as plt
import random
import pandas as pd
import openpyxl
import string
import numpy as np


class agent:
    def __init__(self, position, knowledge, type, spor, pts):
        self.position = position
        self.knowledge = knowledge
        self.type = type
        self.spor = spor
        self.pts = pts


def func(people):
    rus = 0
    hohl = 0
    for i in people:
        if i.type == 1:
            rus += 1
        else:
            hohl += 1
    if rus == 0 or hohl == 0:
        return 1
    rus_dol_year = 11600
    ucr_dol_year = 2724
    T = random.uniform(0, 0.5)
    one_one = 13
    return (rus_dol_year / ucr_dol_year) * (hohl / rus) * T * one_one


def find_topic(x, topic):
    if topic in x:
        return True
    return False


def srach_1(x, y, qwerty, intense):
    wiener = x.type
    flag = False
    num = 0
    while flag == False:
        num = random.randint(2, 28)
        if x.knowledge.cell(row=num, column=3).value != 1 and x.knowledge.cell(row=num, column=3).value is not None:
            flag = True
    wiener = float(x.knowledge.cell(row=num, column=3).value) * y.pts
    stroka = '-----------------Базар' + str(qwerty) + '----------------- \n'
    hohlosrach.write(stroka)
    hohlosrach.write('\n')
    hohlosrach.write(str(num))
    hohlosrach.write('\n')
    hohlosrach.write(str(x.knowledge.cell(row=num, column=2).value))
    hohlosrach.write('\n')
    hohlosrach.write(str(wiener))
    hohlosrach.write('\n')
    flag = False
    while flag == False:
        num = random.randint(2, 28)
        if y.knowledge.cell(row=num, column=5).value != 1 and x.knowledge.cell(row=num, column=3).value is not None:
            flag = True
    wiener = float(y.knowledge.cell(row=num, column=3).value) * y.pts
    hohlosrach.write(str(num))
    hohlosrach.write('\n')
    hohlosrach.write(str(y.knowledge.cell(row=num, column=2).value))
    hohlosrach.write('\n')
    hohlosrach.write(str(wiener))
    hohlosrach.write('\n')
    flag = False
    if float(x.knowledge.cell(row=num, column=3).value) == float(y.knowledge.cell(row=num, column=3).value):
        x.pts += abs(x.position) + 0.01 * intense
        y.pts += abs(y.position) + 0.01 * intense
        hohlosrach.write('123321\n')
    elif float(x.knowledge.cell(row=num, column=3).value) + x.pts > float(
            y.knowledge.cell(row=num, column=3).value) + y.pts:
        x.pts += abs(x.position) + 0.01 * intense
        hohlosrach.write('1\n')
    elif float(x.knowledge.cell(row=num, column=3).value) + x.pts < float(
            y.knowledge.cell(row=num, column=3).value) + y.pts:
        y.pts += abs(y.position) + 0.01 * intense
        hohlosrach.write('2\n')
    if x.pts > 100 and x.position >= 0:
        stroka = 'Хохла порвало\n'
        hohlosrach.write(stroka)
        return 1
    elif y.pts > 100 and x.position >= 0:
        stroka = 'Хохла порвало\n'
        hohlosrach.write(stroka)
        return 1
    elif x.pts > 100 and x.position < 0:
        stroka = 'Слава Хохляндии\n'
        hohlosrach.write(stroka)
        return 2
    elif y.pts > 100 and x.position < 0:
        stroka = 'Слава Хохляндии\n'
        hohlosrach.write(stroka)
        return 2


def srach(x, y, qwerty, hohlosrach, intense):
    wiener = x.type
    flag = False
    num = 0
    if x.type == y.type:
        stroka = '-----------------Базар' + str(qwerty) + '-----------------\n'
        hohlosrach.write(stroka)
        if x.type == 1:
            stroka = 'Русские: ' + str(x.knowledge.cell(row=24, column=2).value)
            hohlosrach.write(stroka)
            hohlosrach.write('\n')
        else:
            stroka = 'Хохлы: ' + str(y.knowledge.cell(row=13, column=2).value)
            hohlosrach.write(stroka)
            hohlosrach.write('\n')
        hohlosrach.write(str(wiener))
        hohlosrach.write('\n')
    else:
        one = random.randint(0, 10)
        two = random.randint(0, 10)
        if one > two:
            wiener = srach_1(x, y, qwerty, intense)
        elif two > one:
            wiener = srach_1(y, x, qwerty, intense)
    return wiener


people = []
russian = openpyxl.load_workbook('knowledge\\Russian.xlsx', data_only=True)
hohli = openpyxl.load_workbook('knowledge\\hohli.xlsx', data_only=True)
hohlosrach = open('knowledge\\hohlosrach.txt', 'w')
hohlosrach.truncate()
russians = russian.active
hohlin = hohli.active

russkie = 0

for i in range(0, 2000):
    pos = random.uniform(-10, 10)
    if 0 <= pos:
        ag = agent(pos, russians, 1, False, 1)
        people.append(ag)
        russkie += 1
    elif pos < 0:
        ag = agent(pos, hohlin, 2, False, 1)
        people.append(ag)
    # people.sort(key=lambda x: x.type)

rus = 0
hohl = 0
for i in people:
    if i.type == 1:
        rus += 1
    else:
        hohl += 1
print('Русские:', rus)
print('Хохлы', hohl)
for srachi in range(0, 50):
    intense = func(people)
    if intense == 1:
        break
    if intense > 15:
        for i in range(2, russians.max_row):
            if russians.cell(row=i, column=4).value == 0:
                russians.cell(row=i, column=4).value = 1
                break
    if intense > 17:
        for i in range(2, hohlin.max_row):
            if hohlin.cell(row=i, column=4).value == 0:
                hohlin.cell(row=i, column=4).value = 1
                break

    np.random.shuffle(people)
    j = len(people) - 1
    stroka = '+++++++++++++++}{0}{0срач №' + str(srachi) + '+++++++++++++++\n'
    hohlosrach.write(stroka)
    for i in range(0, len(people)):
        if j < i:
            break
        if people[i].spor == False and people[j].spor == False:
            wiener = srach(people[i], people[j], i, hohlosrach, intense)
            j -= 1
            if wiener == 1 and people[i].position >= 0:
                #people.pop(j)
                people[j].position = abs(people[j].position)
                people[j].type = 1
                people[j].knowledge = russians
            elif wiener == 1 and people[j].position >= 0:
                #people.pop(i)
                people[i].position = abs(people[i].position)
                people[i].type = 1
                people[i].knowledge = russians
            elif wiener == 2 and people[i].position >= 0:
                #people.pop(j)
                people[j].position = -1 * people[j].position
                people[j].type = 2
                people[j].knowledge = hohlin
            elif wiener == 2 and people[j].position >= 0:
                #people.pop(i)
                people[i].position = -1 * people[i].position
                people[i].type = 2
                people[i].knowledge = hohlin

    for asd in people:
        asd.spor = False
    print('Интенсивность срача: ', intense, '\n')
zxc = 0
for i in people:
    zxc += 1
    print(zxc, 'Кто по масти:', i.type, 'Скил срача', i.pts)

rus = 0
hohl = 0
for i in people:
    if i.type == 1:
        rus += 1
    else:
        hohl += 1
print('Русские:', rus)
print('Хохлы', hohl)