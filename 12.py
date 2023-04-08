import matplotlib.pyplot as plt
import random
import pandas as pd
import openpyxl
import string


class agent:
    def __init__(self, position, knowledge, type, spor):
        self.position = position
        self.knowledge = knowledge
        self.type = type
        self.spor = spor


def start(x):
    if 'старт' in x:
        return True
    return False


def srach(x, y, qwerty):
    wiener = x.type
    flag = False
    num = 0
    if x.type == y.type:
        print('-----------------Базар', qwerty, '-----------------')
        if x.type == 1:
            print(x.knowledge.cell(row=25, column=2).value)
        else:
            print(y.knowledge.cell(row=20, column=2).value)
        print(wiener)
    else:
        one = random.randint(0, 10)
        two = random.randint(0, 10)
        if one > two:
            while flag == False:
                num = random.randint(2, 25)
                negi = x.knowledge.cell(row=num, column=6).value
                ger = x.knowledge.cell(row=num, column=4).value
                if (x.knowledge.cell(row=num, column=6).value == 1) & (
                        start(x.knowledge.cell(row=num, column=4).value) == True) & (
                        x.knowledge.cell(row=num, column=3).value != 1):
                    flag = True
            wiener = x.knowledge.cell(row=num, column=3).value
            print('-----------------Базар', qwerty, '-----------------')
            print(x.knowledge.cell(row=num, column=2).value)
            print(wiener)
        if one < two:
            while flag == False:
                num = random.randint(2, 20)
                negi = y.knowledge.cell(row=num, column=6).value
                ger = y.knowledge.cell(row=num, column=4).value
                rs = start(y.knowledge.cell(row=num, column=4).value)
                if (y.knowledge.cell(row=num, column=6).value == 1) & (
                        start(y.knowledge.cell(row=num, column=4).value) == True):
                    flag = True
            wiener = y.knowledge.cell(row=num, column=3).value
            print('-----------------Базар', qwerty, '-----------------')
            print(y.knowledge.cell(row=num, column=2).value)
            print(wiener)
    return wiener


people = []
russian = openpyxl.load_workbook('knowledge\\Russian.xlsx', data_only=True)
hohli = openpyxl.load_workbook('knowledge\\hohli.xlsx', data_only=True)
russians = russian.active
hohlin = hohli.active

russkie = 0
for i in range(0, 20):
    pos = random.uniform(-1, 1)
    if 0 <= pos:
        ag = agent(pos, russians, 1, False)
        people.append(ag)
        russkie += 1
    elif pos < 0:
        ag = agent(pos, hohlin, 2, False)
        people.append(ag)
people.sort(key=lambda x: x.type)
for i in people:
    print(i.type)
print()
j = len(people) - 1
for i in range(0, len(people)):
    if j < i:
        break
    if people[i].spor == False and people[j].spor == False:
        srach(people[i], people[j], i)
        j -= 1
